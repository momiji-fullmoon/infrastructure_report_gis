import hashlib, math, os, platform, re, time, tracemalloc, unicodedata, zipfile
from datetime import datetime, timezone
from pathlib import Path
from geoalchemy2.elements import WKTElement
from sqlalchemy import insert as sa_insert, select, text
from sqlalchemy.dialects.postgresql import insert as pg_insert
from app.models.models import Pond

IMPORTER_VERSION = "real-excel-dms-v3"
SOURCE_SYSTEM = "tameike_ichiranR8"
REAL_COLUMNS = ["名称","都道府県","市区町村","町域名、番地","緯度","分","秒","経度","分2","秒2","堤高(m)","堤頂長(m)","総貯水量(千m3)","満水面積(km2)"]
COLUMN_MAP = {"name":"名称", "prefecture":"都道府県", "municipality":"市区町村", "address":"町域名、番地", "latitude_degree":"緯度", "latitude_minute":"分", "latitude_second":"秒", "longitude_degree":"経度", "longitude_minute":"分2", "longitude_second":"秒2", "dam_height_m":"堤高(m)", "crest_length_m":"堤頂長(m)", "total_storage_thousand_m3":"総貯水量(千m3)", "full_water_area_km2":"満水面積(km2)"}
SENTINELS = {9999.0, 9999.9}

def norm_text(v):
    if v is None or str(v).strip() == "": return None
    return unicodedata.normalize("NFKC", str(v)).strip()

def _issue(issues, code, field, raw): issues.append({"code": code, "field": field, "rawValue": None if raw is None else str(raw)})
def issue_codes(issues): return [i.get("code") if isinstance(i, dict) else i for i in issues]

def _raw_float(v):
    if v is None or str(v).strip() == "": return None
    s = unicodedata.normalize("NFKC", str(v)).strip().replace(",", "")
    try: return float(s)
    except ValueError: return None

def _parse_number(v, issues, field, sentinel_missing=True):
    f = _raw_float(v)
    if f is None:
        if v is not None and str(v).strip() != "": _issue(issues, "invalid_numeric_value", field, v)
        return None
    if math.isnan(f) or (sentinel_missing and f in SENTINELS):
        _issue(issues, "sentinel_missing_value", field, v); return None
    return f

def parse_degree(v, issues=None, field="degree", kind="latitude"):
    issues = issues if issues is not None else []
    f = _parse_number(v, issues, field)
    if f is None: return None
    lo, hi = (20, 46) if kind == "latitude" else (122, 154)
    if f == 0 or not (lo <= f <= hi): _issue(issues, "invalid_degree", field, v)
    return f

def parse_minute(v, issues=None, field="minute"):
    issues = issues if issues is not None else []
    f = _parse_number(v, issues, field)
    if f is None: return None
    if not (0 <= f < 60): _issue(issues, "invalid_minute", field, v)
    return f

def parse_second(v, issues=None, field="second"):
    issues = issues if issues is not None else []
    f = _parse_number(v, issues, field)
    if f is None: return None
    if not (0 <= f < 60): _issue(issues, "invalid_second", field, v)
    return f

def parse_measurement(v, issues=None, field="measurement", zero_missing=False):
    issues = issues if issues is not None else []
    f = _parse_number(v, issues, field)
    if f == 0 and zero_missing: _issue(issues, "sentinel_missing_value", field, v); return None
    return f

def maybe_float(v, flags=None):
    issues=[]; val=parse_measurement(v, issues)
    if flags is not None:
        for i in issues: flags.add(i["code"])
    return val

def dms_to_decimal(deg, minute=None, second=None, flags=None, kind="latitude"):
    issues = []
    if minute is None and second is None and isinstance(deg, str):
        raw = unicodedata.normalize("NFKC", deg).strip(); nums = [float(x) for x in re.findall(r"-?\d+(?:\.\d+)?", raw)]
        if len(nums) >= 2:
            sign = -1 if nums[0] < 0 else 1; return nums[0] + sign * ((nums[1] if len(nums)>1 else 0)/60 + (nums[2] if len(nums)>2 else 0)/3600)
    d = parse_degree(deg, issues, "latitude_degree" if kind=="latitude" else "longitude_degree", kind)
    m = parse_minute(minute, issues, "latitude_minute" if kind=="latitude" else "longitude_minute")
    s = parse_second(second, issues, "latitude_second" if kind=="latitude" else "longitude_second")
    if flags is not None:
        if isinstance(flags, set): flags.update(i["code"] for i in issues)
        else: flags.extend(issues)
    if d is None or any(i["code"] in {"invalid_degree","invalid_minute","invalid_second","sentinel_missing_value"} for i in issues): return None
    return d + (m or 0) / 60 + (s or 0) / 3600

def sha256(path):
    h=hashlib.sha256()
    with open(path,'rb') as f:
        for b in iter(lambda:f.read(1024*1024),b''): h.update(b)
    return h.hexdigest()

def validate_xlsx(path):
    if Path(path).suffix.lower()!='.xlsx': raise ValueError('input must be .xlsx')
    if os.path.getsize(path)>1024*1024*1024: raise ValueError('input exceeds 1GB safety limit')
    if not zipfile.is_zipfile(path): raise ValueError('xlsx ZIP structure is invalid')

def _norm_name(name): return re.sub(r"(ため池|溜池|池|\s+)", "", unicodedata.normalize("NFKC", name or "").lower())
def _hash(rec):
    keys=["name","prefecture","municipality","address","latitude","longitude","dam_height_m","crest_length_m","total_storage_thousand_m3","full_water_area_km2"]
    return hashlib.sha256(repr([(k, rec.get(k)) for k in keys]).encode()).hexdigest()

def build_record(row, idx, file_hash=""):
    issues=[]; source_payload={c: (str(row.get(c)) if row.get(c) is not None else None) for c in REAL_COLUMNS}; name=norm_text(row.get(COLUMN_MAP['name'])) or f"unknown-{idx}"
    lat=dms_to_decimal(row.get('緯度'), row.get('分'), row.get('秒'), issues, "latitude"); lon=dms_to_decimal(row.get('経度'), row.get('分2'), row.get('秒2'), issues, "longitude")
    area=parse_measurement(row.get('満水面積(km2)'), issues, 'full_water_area_km2')
    if area is None: _issue(issues, 'missing_full_water_area', 'full_water_area_km2', row.get('満水面積(km2)'))
    raw_lat=_raw_float(row.get('緯度')); raw_lon=_raw_float(row.get('経度'))
    if raw_lat is not None and raw_lon is not None and 122 <= raw_lat <= 154 and 20 <= raw_lon <= 46: _issue(issues, 'coordinate_swapped_candidate', 'location', f'{raw_lat},{raw_lon}')
    if lat is None or lon is None: _issue(issues, 'coordinate_missing', 'location', None)
    elif not (20 <= lat <= 46 and 122 <= lon <= 154): _issue(issues, 'coordinate_out_of_japan', 'location', f'{lat},{lon}')
    normalized={'name':name,'latitude':lat,'longitude':lon,'latitude_degree':_raw_float(row.get('緯度')),'latitude_minute':_raw_float(row.get('分')),'latitude_second':_raw_float(row.get('秒')),'longitude_degree':_raw_float(row.get('経度')),'longitude_minute':_raw_float(row.get('分2')),'longitude_second':_raw_float(row.get('秒2')),'input_sha256':file_hash,'importer_version':IMPORTER_VERSION}
    rec={'source_system':SOURCE_SYSTEM,'source_record_id':str(idx),'name':name,'normalized_name':_norm_name(name),'prefecture':norm_text(row.get('都道府県')),'municipality':norm_text(row.get('市区町村')),'municipality_code':None,'address':norm_text(row.get('町域名、番地')),'latitude':lat,'longitude':lon,'coordinate_quality':'map_selected' if lat is not None and lon is not None and 'coordinate_out_of_japan' not in issue_codes(issues) else 'unknown','quality_flags':{'issues':issues},'duplicate_candidate':False,'dam_height_m':parse_measurement(row.get('堤高(m)'), issues, 'dam_height_m'),'crest_length_m':parse_measurement(row.get('堤頂長(m)'), issues, 'crest_length_m'),'total_storage_thousand_m3':parse_measurement(row.get('総貯水量(千m3)'), issues, 'total_storage_thousand_m3'),'full_water_area_km2':area,'source_payload':source_payload,'normalized_payload':normalized,'confidence':0.8,'data_source':'ため池台帳','updated_at':datetime.now(timezone.utc)}
    rec['normalized_payload']['content_hash']=_hash(rec)
    if lon is not None and lat is not None: rec['location']=WKTElement(f"POINT({lon} {lat})", srid=4326)
    else: rec['location']=None
    return rec

def _existing_hashes(db, batch):
    keys=[r['source_record_id'] for r in batch]
    rows=db.execute(select(Pond.source_record_id, Pond.normalized_payload).where(Pond.source_system==SOURCE_SYSTEM, Pond.source_record_id.in_(keys))).all()
    return {rid: (payload or {}).get('content_hash') for rid,payload in rows}

def flush(db, batch, mode):
    if not batch or mode=='dry-run': return {'inserted':0,'updated':0,'skipped':len(batch) if mode=='dry-run' else 0}
    existing=_existing_hashes(db,batch); new=[r for r in batch if r['source_record_id'] not in existing]; changed=[r for r in batch if r['source_record_id'] in existing and existing[r['source_record_id']] != r['normalized_payload'].get('content_hash')]
    skipped=len(batch)-len(new)-len(changed); todo=new+changed
    if not todo: db.commit(); return {'inserted':0,'updated':0,'skipped':skipped}
    if db.bind.dialect.name=='postgresql':
        stmt=pg_insert(Pond).values(todo)
        if mode=='insert': stmt=stmt.on_conflict_do_nothing(index_elements=['source_system','source_record_id'])
        else: stmt=stmt.on_conflict_do_update(index_elements=['source_system','source_record_id'], set_={k:getattr(stmt.excluded,k) for k in todo[0] if k not in ['source_system','source_record_id','pond_id','created_at']})
        db.execute(stmt)
    else:
        if mode == 'insert': todo=new

        if todo:
            sqlite_todo=[{k:v for k,v in r.items() if k!='location'} for r in todo]
            db.execute(sa_insert(Pond), sqlite_todo)
    db.commit(); return {'inserted':len(new), 'updated':0 if mode=='insert' else len(changed), 'skipped':skipped + (len(changed) if mode=='insert' else 0)}

def import_excel(db, input_path, limit=None, mode='upsert', batch_size=5000, fail_after_read=False):
    from openpyxl import load_workbook
    if mode not in {'insert','upsert','replace-source','dry-run'}: raise ValueError('unsupported import mode')
    validate_xlsx(input_path); start=time.perf_counter(); tracemalloc.start(); file_hash=sha256(input_path)
    wb=load_workbook(input_path, read_only=True, data_only=True); ws=wb[wb.sheetnames[0]]; headers=[str(c).strip() if c is not None else f'col{i}' for i,c in enumerate(next(ws.iter_rows(values_only=True)))]
    missing=[c for c in REAL_COLUMNS if c not in headers]
    if missing: raise ValueError(f'missing real Excel columns: {missing}')
    stats={'input':str(input_path),'sha256':file_hash,'sheet':ws.title,'headers':headers,'mode':mode,'batchSize':batch_size,'processed':0,'inserted':0,'updated':0,'skipped':0,'failed':0,'missingCoordinates':0,'outOfJapan':0,'duplicateExactCoordinate':0,'duplicateRoundedCoordinate':0,'duplicateSimilarName':0,'duplicateNearbyCandidate':0,'failureSamples':[],'importerVersion':IMPORTER_VERSION,'platform':platform.platform()}
    batch=[]; all_records=[]; seen_exact=set(); seen_round=set(); seen_names={}
    for idx, vals in enumerate(ws.iter_rows(values_only=True), start=2):
        if limit and stats['processed']>=limit: break
        stats['processed']+=1
        try:
            rec=build_record(dict(zip(headers, vals)), idx, file_hash); codes=issue_codes(rec['quality_flags']['issues'])
            if 'coordinate_missing' in codes: stats['missingCoordinates']+=1
            if 'coordinate_out_of_japan' in codes: stats['outOfJapan']+=1
            if rec['latitude'] is not None and rec['longitude'] is not None:
                exact=(rec['latitude'],rec['longitude']); rounded=(round(rec['latitude'],6),round(rec['longitude'],6)); mun=(rec['prefecture'],rec['municipality'])
                if exact in seen_exact: rec['quality_flags']['issues'].append({'code':'duplicate_exact_coordinate','field':'location','rawValue':str(exact)}); stats['duplicateExactCoordinate']+=1
                if rounded in seen_round: rec['quality_flags']['issues'].append({'code':'duplicate_rounded_coordinate','field':'location','rawValue':str(rounded)}); stats['duplicateRoundedCoordinate']+=1
                seen_exact.add(exact); seen_round.add(rounded); key=(mun,rec['normalized_name'])
                if key in seen_names: rec['quality_flags']['issues'].append({'code':'duplicate_similar_name','field':'normalized_name','rawValue':rec['name']}); stats['duplicateSimilarName']+=1
                seen_names[key]=1
            rec['duplicate_candidate']=any(str(i.get('code') if isinstance(i,dict) else i).startswith('duplicate_') for i in rec['quality_flags']['issues'])
            if mode=='replace-source': all_records.append(rec)
            else:
                batch.append(rec)
                if len(batch)>=batch_size:
                    r=flush(db,batch,mode); [stats.__setitem__(k, stats[k]+r[k]) for k in r]; batch=[]
        except Exception as e:
            stats['failed']+=1
            if len(stats['failureSamples'])<100: stats['failureSamples'].append({'row':idx,'error':type(e).__name__,'message':str(e)})
    if fail_after_read: raise RuntimeError('intentional import failure after read')
    if mode=='replace-source':
        with db.begin():
            db.execute(text("delete from pond where source_system=:s"), {'s':SOURCE_SYSTEM})
            rows=all_records if db.bind.dialect.name=='postgresql' else [{k:v for k,v in r.items() if k!='location'} for r in all_records]
            if rows: db.execute(sa_insert(Pond), rows)
        stats['inserted']=len(all_records)
    else:
        r=flush(db,batch,mode); [stats.__setitem__(k, stats[k]+r[k]) for k in r]
    cur,peak=tracemalloc.get_traced_memory(); tracemalloc.stop(); elapsed=time.perf_counter()-start
    dbsize=None
    if db.bind.dialect.name=='postgresql':
        try: dbsize=db.scalar(text('select pg_database_size(current_database())'))
        except Exception: dbsize=None
    stats.update({'elapsedSeconds':round(elapsed,3),'rowsPerSecond':round(stats['processed']/elapsed,2) if elapsed else 0,'peakMemoryBytes':peak,'databaseSizeBytes':dbsize,'completedAt':datetime.now(timezone.utc).isoformat()})
    return stats
